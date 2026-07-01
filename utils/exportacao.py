from __future__ import annotations

import csv
import os
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill


_COR_CABECALHO = "9F3FFA"
_COR_ALERTA = "FFCCCC"


def exportar_excel(
    caminho: str,
    titulo_aba: str,
    cabecalhos: list[str],
    linhas: list[tuple[Any, ...]],
    colunas_alerta: dict[int, Any] | None = None,
) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = titulo_aba

    fonte_cabecalho = Font(bold=True, color="FFFFFF", size=11)
    fill_cabecalho = PatternFill("solid", fgColor=_COR_CABECALHO)
    alinhamento_centro = Alignment(horizontal="center", vertical="center")

    for col_idx, cabecalho in enumerate(cabecalhos, start=1):
        cell = ws.cell(row=1, column=col_idx, value=cabecalho)
        cell.font = fonte_cabecalho
        cell.fill = fill_cabecalho
        cell.alignment = alinhamento_centro

    for row_idx, linha in enumerate(linhas, start=2):
        for col_idx, valor in enumerate(linha, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.alignment = alinhamento_centro
            if colunas_alerta and col_idx in colunas_alerta:
                col_ref, col_alerta = colunas_alerta[col_idx]
                val_ref = linha[col_ref]
                val_alerta = linha[col_alerta]
                if isinstance(val_ref, (int, float)) and isinstance(val_alerta, (int, float)):
                    if val_ref <= val_alerta:
                        for c in range(1, len(linha) + 1):
                            ws.cell(row=row_idx, column=c).fill = PatternFill(
                                "solid", fgColor=_COR_ALERTA
                            )

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    wb.save(caminho)


def exportar_csv(
    caminho: str,
    cabecalhos: list[str],
    linhas: list[tuple[Any, ...]],
) -> None:
    with open(caminho, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f, delimiter=";")
        writer.writerow(cabecalhos)
        writer.writerows(linhas)


def importar_excel(caminho: str, linha_cabecalho: int = 1) -> tuple[list[str], list[list[Any]]]:
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []

    cabecalhos = [str(c) if c is not None else "" for c in rows[linha_cabecalho - 1]]
    dados = [
        [cell if cell is not None else "" for cell in row]
        for row in rows[linha_cabecalho:]
        if any(cell is not None for cell in row)
    ]
    return cabecalhos, dados


def importar_csv(caminho: str) -> tuple[list[str], list[list[Any]]]:
    encoding = _detectar_encoding(caminho)
    with open(caminho, newline="", encoding=encoding) as f:
        reader = csv.reader(f, delimiter=_detectar_delimitador(caminho, encoding))
        rows = list(reader)

    if not rows:
        return [], []

    return rows[0], rows[1:]


def _detectar_encoding(caminho: str) -> str:
    for enc in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
        try:
            with open(caminho, encoding=enc) as f:
                f.read()
            return enc
        except UnicodeDecodeError:
            continue
    return "latin-1"


def _detectar_delimitador(caminho: str, encoding: str) -> str:
    with open(caminho, encoding=encoding) as f:
        amostra = f.read(2048)
    contagens = {d: amostra.count(d) for d in (";", ",", "\t", "|")}
    return max(contagens, key=contagens.get)
